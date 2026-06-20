"""
Export engine — generate PDF and Excel reports from analytics data.

Supports Arabic content with proper RTL rendering for PDF exports.
"""

import io
from enum import Enum
from typing import Any

import structlog

logger = structlog.get_logger(__name__)


class ExportFormat(Enum):
    """Supported export formats."""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    JSON = "json"


class ExportEngine:
    """
    Export engine for generating reports in multiple formats.

    Supports PDF (with Arabic RTL), Excel, CSV, and JSON exports.
    """

    def __init__(self) -> None:
        self._initialized = False

    def export(
        self,
        data: dict[str, Any],
        format: ExportFormat,
        title: str = "Report",
        filename: str | None = None,
    ) -> io.BytesIO:
        """
        Export data to the specified format.

        Args:
            data: Report data dictionary
            format: Export format
            title: Report title (for PDF)
            filename: Optional custom filename

        Returns:
            BytesIO object containing the exported file
        """
        if format == ExportFormat.PDF:
            return self._export_pdf(data, title)
        elif format == ExportFormat.EXCEL:
            return self._export_excel(data, title)
        elif format == ExportFormat.CSV:
            return self._export_csv(data)
        elif format == ExportFormat.JSON:
            return self._export_json(data)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _export_pdf(self, data: dict[str, Any], title: str) -> io.BytesIO:
        """Export data as PDF with Arabic support."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.units import inch, cm
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm)

            # Try to register Arabic font
            try:
                pdfmetrics.registerFont(TTFont('Arabic', 'fonts/arial.ttf'))
                font_name = 'Arabic'
            except Exception:
                font_name = 'Helvetica'

            styles = getSampleStyleSheet()

            # Custom title style for Arabic
            title_style = ParagraphStyle(
                'ArabicTitle',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=18,
                alignment=2,  # Right alignment for Arabic
                spaceAfter=20,
            )

            body_style = ParagraphStyle(
                'ArabicBody',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=12,
                alignment=2,  # Right alignment
                spaceAfter=10,
            )

            elements = []

            # Title
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 0.5*inch))

            # Content
            for key, value in data.items():
                if isinstance(value, dict):
                    # Nested dictionary - create table
                    table_data = [[k, str(v)] for k, v in value.items()]
                    if table_data:
                        table = Table(table_data, colWidths=[4*inch, 4*inch])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                            ('FONTNAME', (0, 0), (-1, -1), font_name),
                            ('FONTSIZE', (0, 0), (-1, -1), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ]))
                        elements.append(Paragraph(str(key), body_style))
                        elements.append(table)
                        elements.append(Spacer(1, 0.3*inch))
                elif isinstance(value, list):
                    # List - bullet points
                    elements.append(Paragraph(f"<b>{key}:</b>", body_style))
                    for item in value:
                        elements.append(Paragraph(f"• {item}", body_style))
                else:
                    # Simple key-value
                    elements.append(Paragraph(f"<b>{key}:</b> {value}", body_style))

            doc.build(elements)
            buffer.seek(0)
            logger.info("export.pdf_generated", title=title, size=buffer.getbuffer().nbytes)
            return buffer

        except ImportError:
            # Fallback: return text-based PDF
            return self._export_pdf_fallback(data, title)

    def _export_pdf_fallback(self, data: dict[str, Any], title: str) -> io.BytesIO:
        """Fallback PDF export using basic text."""
        buffer = io.BytesIO()

        # Create a simple text file as fallback
        content = f"{title}\n{'='*50}\n\n"
        for key, value in data.items():
            if isinstance(value, dict):
                content += f"\n{key}:\n"
                for k, v in value.items():
                    content += f"  {k}: {v}\n"
            elif isinstance(value, list):
                content += f"\n{key}:\n"
                for item in value:
                    content += f"  • {item}\n"
            else:
                content += f"{key}: {value}\n"

        buffer.write(content.encode('utf-8'))
        buffer.seek(0)
        return buffer

    def _export_excel(self, data: dict[str, Any], title: str) -> io.BytesIO:
        """Export data as Excel workbook."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title[:31]  # Excel sheet name max 31 chars

            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'),
            )

            # Title row
            ws.merge_cells('A1:D1')
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='right')

            row = 3
            for key, value in data.items():
                if isinstance(value, dict):
                    # Header for section
                    ws.cell(row=row, column=1, value=key)
                    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                    row += 1

                    # Table headers
                    headers = list(value.keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                    row += 1

                    # Table data
                    for v in value.values():
                        ws.cell(row=row, column=1, value=str(v)).border = border
                        row += 1

                    row += 1  # Spacing
                elif isinstance(value, list):
                    ws.cell(row=row, column=1, value=key)
                    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                    row += 1
                    for item in value:
                        ws.cell(row=row, column=1, value=f"• {item}")
                        row += 1
                    row += 1
                else:
                    ws.cell(row=row, column=1, value=key)
                    ws.cell(row=row, column=1).font = Font(bold=True)
                    ws.cell(row=row, column=2, value=str(value))
                    row += 1

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            logger.info("export.excel_generated", title=title, size=buffer.getbuffer().nbytes)
            return buffer

        except ImportError:
            return self._export_csv(data)

    def _export_csv(self, data: dict[str, Any]) -> io.BytesIO:
        """Export data as CSV with UTF-8 BOM for Excel compatibility."""
        import csv

        # Use TextIOWrapper to add BOM for Excel
        text_buffer = io.StringIO()
        text_buffer.write('\ufeff')  # UTF-8 BOM
        writer = csv.writer(text_buffer)

        for key, value in data.items():
            if isinstance(value, dict):
                writer.writerow([key])
                for k, v in value.items():
                    writer.writerow([k, str(v)])
                writer.writerow([])
            elif isinstance(value, list):
                writer.writerow([key])
                for item in value:
                    writer.writerow([str(item)])
                writer.writerow([])
            else:
                writer.writerow([key, str(value)])

        # Convert to BytesIO
        buffer = io.BytesIO()
        buffer.write(text_buffer.getvalue().encode('utf-8'))
        buffer.seek(0)
        return buffer

    def _export_json(self, data: dict[str, Any]) -> io.BytesIO:
        """Export data as JSON."""
        import json

        buffer = io.BytesIO()
        content = json.dumps(data, ensure_ascii=False, indent=2)
        buffer.write(content.encode('utf-8'))
        buffer.seek(0)
        return buffer

    def get_export_content_type(self, format: ExportFormat) -> str:
        """Get MIME content type for format."""
        content_types = {
            ExportFormat.PDF: "application/pdf",
            ExportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ExportFormat.CSV: "text/csv; charset=utf-8",
            ExportFormat.JSON: "application/json; charset=utf-8",
        }
        return content_types.get(format, "application/octet-stream")

    def get_export_extension(self, format: ExportFormat) -> str:
        """Get file extension for format."""
        extensions = {
            ExportFormat.PDF: ".pdf",
            ExportFormat.EXCEL: ".xlsx",
            ExportFormat.CSV: ".csv",
            ExportFormat.JSON: ".json",
        }
        return extensions.get(format, ".bin")


# Global export engine
export_engine = ExportEngine()
