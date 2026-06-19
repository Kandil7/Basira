"""
Document text extraction utilities for PDF and Excel files.

Provides text extraction from common document formats for the
Internal Ops Agent document upload pipeline.
"""

import io
import structlog

logger = structlog.get_logger(__name__)


def extract_text_from_pdf(content: bytes) -> str:
    """
    Extract text from PDF bytes using pypdf.

    Args:
        content: Raw PDF file bytes

    Returns:
        Extracted text with page separators.

    Raises:
        ValueError: If PDF is empty or corrupted
    """
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(content))
        text_parts: list[str] = []

        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text and text.strip():
                text_parts.append(f"--- Page {i + 1} ---\n{text.strip()}")

        if not text_parts:
            raise ValueError("PDF contains no extractable text (may be image-only)")

        result = "\n\n".join(text_parts)
        logger.info(
            "extractor.pdf_success",
            pages=len(reader.pages),
            text_length=len(result),
        )
        return result

    except ImportError:
        logger.error("extractor.pdf_missing_dep", dep="pypdf")
        raise ValueError("PDF extraction not available — pypdf not installed")
    except Exception as e:
        if "ValueError" in type(e).__name__:
            raise
        logger.error("extractor.pdf_failed", error=str(e))
        raise ValueError(f"Failed to extract PDF text: {e}")


def extract_text_from_excel(content: bytes) -> str:
    """
    Extract text from Excel bytes using openpyxl.

    Reads all sheets and formats them as structured text.

    Args:
        content: Raw Excel file bytes (.xlsx)

    Returns:
        Extracted text with sheet headers and pipe-separated rows.

    Raises:
        ValueError: If Excel is empty or corrupted
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        text_parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_lines: list[str] = [f"=== Sheet: {sheet_name} ==="]

            row_count = 0
            for row in ws.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(v.strip() for v in row_values):
                    sheet_lines.append(" | ".join(row_values))
                    row_count += 1

            if row_count > 0:
                text_parts.append("\n".join(sheet_lines))

        wb.close()

        if not text_parts:
            raise ValueError("Excel file contains no data")

        result = "\n\n".join(text_parts)
        logger.info(
            "extractor.excel_success",
            sheets=len(text_parts),
            text_length=len(result),
        )
        return result

    except ImportError:
        logger.error("extractor.excel_missing_dep", dep="openpyxl")
        raise ValueError("Excel extraction not available — openpyxl not installed")
    except Exception as e:
        if "ValueError" in type(e).__name__:
            raise
        logger.error("extractor.excel_failed", error=str(e))
        raise ValueError(f"Failed to extract Excel text: {e}")


def extract_text_from_csv(content: bytes) -> str:
    """
    Extract text from CSV bytes.

    Args:
        content: Raw CSV file bytes

    Returns:
        Formatted CSV text.
    """
    try:
        text = content.decode("utf-8", errors="replace")
        lines = text.strip().split("\n")
        if not lines:
            raise ValueError("CSV file is empty")
        logger.info("extractor.csv_success", lines=len(lines))
        return text
    except Exception as e:
        logger.error("extractor.csv_failed", error=str(e))
        raise ValueError(f"Failed to extract CSV text: {e}")
